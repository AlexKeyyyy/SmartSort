from __future__ import annotations

from pathlib import Path

TEXT_EXTENSIONS = {
    ".txt",
    ".py",
    ".js",
    ".ts",
    ".java",
    ".cpp",
    ".go",
    ".sql",
    ".md",
    ".log",
    ".csv",
    ".json",
    ".yaml",
    ".yml",
    ".html",
    ".css",
    ".sh",
    ".env",
    ".toml",
}


def _truncate_text(value: str, max_chars: int) -> str:
    return value[:max_chars].strip()


def _read_text_file(path: Path, max_chars: int) -> str:
    with path.open("r", encoding="utf-8", errors="ignore") as file:
        return _truncate_text(file.read(), max_chars)


def _read_pdf(path: Path, max_chars: int) -> str:
    import fitz

    parts: list[str] = []
    total = 0

    with fitz.open(path) as document:
        for page in document:
            page_text = page.get_text("text").strip()
            if not page_text:
                continue

            remaining = max_chars - total
            if remaining <= 0:
                break

            snippet = page_text[:remaining]
            parts.append(snippet)
            total += len(snippet)

    return " ".join(parts).strip()


def _read_docx(path: Path, max_chars: int) -> str:
    from docx import Document

    document = Document(path)
    text = " ".join(paragraph.text.strip() for paragraph in document.paragraphs if paragraph.text.strip())
    return _truncate_text(text, max_chars)


def _read_xlsx(path: Path, max_chars: int) -> str:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    cells: list[str] = []
    total = 0

    try:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                for value in row:
                    if value is None:
                        continue

                    text = str(value).strip()
                    if not text:
                        continue

                    remaining = max_chars - total
                    if remaining <= 0:
                        return " ".join(cells).strip()

                    snippet = text[:remaining]
                    cells.append(snippet)
                    total += len(snippet) + 1
    finally:
        workbook.close()

    return " ".join(cells).strip()


def extract_text(filepath: str | Path, max_chars: int = 500) -> str:
    try:
        path = Path(filepath)
        if max_chars <= 0 or not path.exists() or not path.is_file():
            return ""

        suffix = path.suffix.lower()

        if suffix in TEXT_EXTENSIONS:
            return _read_text_file(path, max_chars)
        if suffix == ".pdf":
            return _read_pdf(path, max_chars)
        if suffix == ".docx":
            return _read_docx(path, max_chars)
        if suffix == ".xlsx":
            return _read_xlsx(path, max_chars)

        return ""
    except Exception:
        return ""
